from app import create_app, db
from app.models import Discipline, Teacher, Modulus, Cohort, Lecture, Classroom
import openpyxl
import json


def clear_all_tables():
    app = create_app()
    with app.app_context():
        db.session.query(Modulus).delete()
        db.session.query(Discipline).delete()
        db.session.query(Cohort).delete()
        db.session.query(Teacher).delete()
        db.session.commit()


def complete_sheet():
    file_name = "db_info/projeto2024.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    
    with open("db_info/disciplines_to_load_.json", "r") as file:
        disciplines_dict = json.load(file)
    
    for row in sheet.iter_rows(min_row=2):
        discipline_code = row[0].value
        row_dict = disciplines_dict.get(discipline_code)
        if not row_dict:
            continue
        row[2].value = row_dict["name_abbr"]
        row[8].value = row_dict["is_theoretical"]
        row[9].value = row_dict["is_intensive"]
    
    workbook.save("db_info/projeto2024_.xlsx")


def convert_sheet():
    file_name = "db_info/projeto2024.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        stripped_row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
        data.append(dict(zip(header, stripped_row)))

    return data


def add_teachers():
    with open("db_info/teachers_to_load.txt", "r") as file:
        teachers = file.readlines()
    
    teachers = [teacher.strip() for teacher in teachers]

    for teacher_name_abbr_rg in teachers:
        if teacher_name_abbr_rg == "":
            continue
        name, abbr, rg = teacher_name_abbr_rg.split(", ")
        if rg == "0":
            continue
        
        name = ' '.join([name.capitalize().strip() for name in name.split(' ')])
        abbr = ' '.join([name.capitalize().strip() for name in abbr.split(' ')])
        rg = rg.replace(".", "").replace("-", "")

        app = create_app()
        with app.app_context():
            check_teacher = Teacher.query.filter_by(rg=rg).first()
            if check_teacher:
                print(f"Teacher {name} already exists.")
                continue

            teacher_dict = {
                "rg": rg,
                "abbr_name": abbr,
                "name": name
            }

            print(teacher_dict)
            Teacher.add_teacher(**teacher_dict)


def add_cohorts():
    app = create_app()
    with app.app_context():
        from db_info.cohorts_to_load import cohorts_to_load
        for cohort in cohorts_to_load:
            print(cohort)
            Cohort.add_cohort(**cohort)


def add_classrooms():
    from db_info.classrooms_to_load import classrooms_to_load
    
    app = create_app()
    with app.app_context():
        for classroom in classrooms_to_load:
            print(classroom)
            Classroom.add_classroom(**classroom)


def add_disciplines():
    disciplines = convert_sheet()

    for discipline in disciplines:
        discipline.pop("observations")
        discipline.pop("cohorts")
        discipline.pop("prerequisites")

        if discipline["workload"] is None:
            discipline["workload"] = 0
        if discipline["teachers"] is not None:
            discipline["teachers"] = discipline["teachers"].split(", ")

        if discipline["is_theoretical"] == '=TRUE()' or discipline["is_theoretical"] == 1:
            discipline["is_theoretical"] = True
        else:
            discipline["is_theoretical"] = False

        if discipline["is_intensive"] == '=TRUE()' or discipline["is_intensive"] == 1:
            discipline["is_intensive"] = True
        else:
            discipline["is_intensive"] = False
        
        if discipline["is_theoretical"]:
            discipline["mandatory_room"] = -1
        
        if discipline["mandatory_room"] is None:
            discipline["mandatory_room"] = 0
        
        if discipline["name_abbr"] is None:
            discipline["name_abbr"] = discipline["name"]

        if discipline["joined_cohorts"] is None:
            discipline["joined_cohorts"] = False
        else:
            discipline["joined_cohorts"] = bool(discipline["joined_cohorts"])

        print(discipline)
        app = create_app()
        with app.app_context():
            Discipline.add_discipline(**discipline)        


def add_prerequisites():
    data = convert_sheet()
    for datum in data:
        code = datum["code"]
        prerequisites = datum["prerequisites"]

        if prerequisites is None:
            continue

        prerequisites = [p.strip() for p in prerequisites.split(",")]
        app = create_app()
        with app.app_context():
            discipline = Discipline.query.filter_by(code=code).first()
            if not discipline:
                print(f"Discipline {code} not found.")
                continue
            
            print(code, prerequisites)

            for prerequisite in prerequisites:
                discipline.add_prerequisite(prerequisite)


def add_moduli():
    data = convert_sheet()
    for datum in data:
        disc_code = datum["code"]
        cohorts_abr = datum["cohorts"]
        
        if cohorts_abr is None:
            print(f"Discipline {disc_code} has no cohorts.")
            continue

        app = create_app()
        with app.app_context():
            discipline = Discipline.query.filter_by(code=disc_code).first()
            if not discipline:
                print(f"Discipline {disc_code} not found.")
                continue
            
            all_cohorts_codes = [c.code for c in Cohort.query.all()]
            cohorts_abrs = [c.strip() for c in cohorts_abr.split("+")]
            # cohort_codes = []
            for cohort_abr in cohorts_abrs:
                cohort_codes = [code for code in all_cohorts_codes if cohort_abr in code]
                for code in cohort_codes:
                    cohort = Cohort.query.filter_by(code=code).first()
                    if not cohort:
                        print(f"Cohort {code} not found.")
                        continue

                    module = Modulus.add_modulus(cohort.code, discipline.code)
                    module.set_main_classroom()
                    module.set_joined_cohorts()
                    print(cohort.code, discipline.code)


def gen_teacher_moduli_dict():
    app = create_app()
    with app.app_context():
        moduli = Modulus.query.all()

        teachers_to_moduli = {}
        for modulus in moduli:
            if modulus.code not in teachers_to_moduli:
                teachers_to_moduli[modulus.code] = []

            for teacher in modulus.discipline.teachers:
                teachers_to_moduli[modulus.code].append(teacher.name)    
            
        with open("db_info/teachers_to_moduli.json", "w") as file:
            json.dump(teachers_to_moduli, file, indent=4)
    

def add_teacher_to_modulus():
    with open("db_info/teachers_to_moduli.json", "r") as file:
        teachers_to_moduli = json.load(file)

    app = create_app()
    with app.app_context():
        for modulus_code, teachers_names in teachers_to_moduli.items():
            disc_code, cohort_code = modulus_code.split("-")
    
            discipline = Discipline.query.filter_by(code=disc_code).first()
            if not discipline:
                print(f"Discipline {disc_code} not found.")
                continue
            
            cohort = Cohort.query.filter_by(code=cohort_code).first()
            if not cohort:
                print(f"Cohort {cohort_code} not found.")
                continue
            
            modulus = Modulus.query.filter_by(discipline_id=discipline.id, cohort_id=cohort.id).first()
            if not modulus:
                print(f"Modulus {modulus_code} not found.")
                continue
            
            for teacher_name in teachers_names:
                teacher_name = ' '.join([name.capitalize().strip() for name in teacher_name.split(' ')])
    
                teacher = Teacher.query.filter_by(name=teacher_name).first()
                if not teacher:
                    print(f"Teacher {teacher_name} not found.")
                    continue
                
                flag = modulus.add_teacher(teacher_name)
                print(flag)


def add_users():
    from app.models import User
    
    app = create_app()
    with app.app_context():
        users =[
            {"name": "admin", "password": "root", "is_admin": True},
        ]

        for user in users:
            User.add_entry(**user)
